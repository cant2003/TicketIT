import os
import sqlite3
import io
import smtplib
from datetime import datetime
from pathlib import Path
from email.message import EmailMessage

import atexit
import sys

from flask import (
    Flask,
    render_template,
    request,
    redirect,
    url_for,
    session,
    jsonify,
    send_file,
)

from dotenv import load_dotenv
from openpyxl import load_workbook
import requests


from services.launcher_service import (
    get_status,
    start_service,
    stop_service,
    start_all,
    stop_all,
    clear_logs,
)


BASE_DIR = Path(__file__).resolve().parent
ROOT_DIR = BASE_DIR.parent


sys.path.append(str(ROOT_DIR))

load_dotenv(ROOT_DIR / ".env")

app = Flask(__name__)

clear_logs()

app.secret_key = os.getenv("SECRET_KEY", "cambia-esta-clave")

def shutdown_services_on_exit():
    try:
        stop_all()
        print("Servicios detenidos automáticamente al cerrar el panel.")
    except Exception as e:
        print(f"No se pudieron detener los servicios: {e}")


atexit.register(shutdown_services_on_exit)

DB_PATH = Path(os.getenv("TICKETIT_DB", str(BASE_DIR / "tickets.db")))

TICKET_TABLE_HINTS = ["tickets", "ticket", "incidencias", "solicitudes"]

USER_TABLE_HINTS = [
    "usuarios_ti",
    "ti_users",
    "usuarios",
    "users",
    "admins",
    "autorizados",
]


def q(name):
    return '"' + str(name).replace('"', '""') + '"'


def con():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c


def tables():
    try:
        with con() as c:
            return [
                r[0]
                for r in c.execute(
                    "select name from sqlite_master where type='table' order by name"
                )
            ]
    except Exception:
        return []


def cols(table):
    with con() as c:
        return [r["name"] for r in c.execute(f"pragma table_info({q(table)})")]


def pick_table(hints):
    ts = tables()
    lower = {t.lower(): t for t in ts}

    for h in hints:
        if h in lower:
            return lower[h]

    for t in ts:
        if any(h in t.lower() for h in hints):
            return t

    return ts[0] if ts else None


def ticket_table():
    return pick_table(TICKET_TABLE_HINTS)


def user_table():
    return pick_table(USER_TABLE_HINTS)


def find_col(names, possible):
    low = {c.lower(): c for c in names}

    for p in possible:
        if p.lower() in low:
            return low[p.lower()]

    for c in names:
        if any(p.lower() in c.lower() for p in possible):
            return c

    return None


def status_class(value):
    s = str(value or "").lower()

    if any(x in s for x in ["abiert", "open", "nuevo"]):
        return "abierto"

    if any(x in s for x in ["proceso", "pend"]):
        return "proceso"

    if any(x in s for x in ["cerr", "closed", "resuelto", "cancel"]):
        return "cerrado"

    return ""


def require_login():
    return "telegram_id" in session


@app.before_request
def guard():
    if request.endpoint in ["login", "static"]:
        return

    if request.path.startswith("/api/") and request.endpoint == "api_login_status":
        return

    if not require_login():
        return redirect(url_for("login"))


@app.route("/", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        telegram_id = request.form.get("telegram_id", "").strip()
        ok, username = validate_telegram_id(telegram_id)

        if ok:
            session["telegram_id"] = telegram_id
            session["username"] = username or f"TI {telegram_id}"
            return redirect(url_for("home"))

        error = "ID de Telegram no autorizado"

    return render_template("login.html", error=error)


def validate_telegram_id(tid):
    if not tid or not DB_PATH.exists():
        return False, None

    try:
        ut = user_table()
        cs = cols(ut) if ut else []

        idc = find_col(
            cs,
            ["telegram_id", "chat_id", "id_telegram", "telegram", "user_id"],
        )

        namec = find_col(cs, ["nombre", "name", "usuario", "username"])

        if not idc:
            return False, None

        with con() as c:
            row = c.execute(
                f"select * from {q(ut)} where cast({q(idc)} as text)=? limit 1",
                (tid,),
            ).fetchone()

            return bool(row), (row[namec] if row and namec else None)

    except Exception:
        return False, None


@app.route("/home")
def home():
    return render_template("home.html", user=session.get("username"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/tickets")
def tickets():
    return render_template(
        "tickets.html",
        table=ticket_table(),
        db_path=str(DB_PATH),
    )


@app.route("/gestion-ti")
def gestion_ti():
    return render_template(
        "gestion_ti.html",
        table=user_table(),
        db_path=str(DB_PATH),
    )


@app.route("/launcher")
def launcher():
    return render_template("launcher.html")


@app.route("/api/stats")
def api_stats():
    out = {
        "tickets_new": 0,
        "total": 0,
        "abiertos": 0,
        "proceso": 0,
        "cerrados": 0,
    }

    tt = ticket_table()

    if not tt:
        return jsonify(out)

    try:
        cs = cols(tt)
        state = find_col(cs, ["estado", "status", "state"])

        with con() as c:
            out["total"] = c.execute(
                f"select count(*) from {q(tt)}"
            ).fetchone()[0]

            if state:
                rows = c.execute(
                    f"""
                    select lower(cast({q(state)} as text)) s, count(*) n
                    from {q(tt)}
                    group by s
                    """
                ).fetchall()

                for r in rows:
                    cl = status_class(r["s"])

                    if cl == "abierto":
                        out["abiertos"] += r["n"]
                    elif cl == "proceso":
                        out["proceso"] += r["n"]
                    elif cl == "cerrado":
                        out["cerrados"] += r["n"]

                out["tickets_new"] = out["abiertos"]

    except Exception as e:
        out["error_msg"] = str(e)

    return jsonify(out)


@app.route("/api/table/<kind>")
def api_table(kind):
    table = ticket_table() if kind == "tickets" else user_table()

    if not table:
        return jsonify({"columns": [], "rows": []})

    search = request.args.get("search", "").strip().lower()
    status = request.args.get("status", "").strip().lower()

    try:
        cs = cols(table)
        stc = find_col(cs, ["estado", "status", "state"])

        with con() as c:
            rows = [
                dict(r)
                for r in c.execute(
                    f"""
                    select rowid as _rowid, *
                    from {q(table)}
                    order by rowid desc
                    limit 500
                    """
                ).fetchall()
            ]

        if search:
            rows = [
                r
                for r in rows
                if search
                in " ".join(
                    str(v or "") for k, v in r.items() if k != "_rowid"
                ).lower()
            ]

        if status and stc:
            rows = [r for r in rows if status_class(r.get(stc)) == status]

        return jsonify({"columns": cs, "rows": rows})

    except Exception as e:
        return jsonify({"columns": ["Error"], "rows": [{"Error": str(e)}]})

@app.get("/api/ti-users")
def api_ti_users():
    ut = user_table()

    if not ut:
        return jsonify([])

    cs = cols(ut)
    idc = find_col(cs, ["telegram_id", "chat_id", "id_telegram", "telegram", "user_id"])
    namec = find_col(cs, ["nombre", "name", "usuario", "username"])

    if not idc:
        return jsonify([])

    with con() as c:
        rows = c.execute(f"select * from {q(ut)} order by {q(namec or idc)}").fetchall()

    return jsonify([
        {
            "telegram_id": str(r[idc]),
            "nombre": r[namec] if namec else str(r[idc]),
        }
        for r in rows
    ])


def enviar_mensaje_telegram(chat_id, texto):
    token = os.getenv("TELEGRAM_TOKEN")

    if not token:
        print("No existe TELEGRAM_TOKEN en .env")
        return False

    url = f"https://api.telegram.org/bot{token}/sendMessage"

    try:
        response = requests.post(
            url,
            json={
                "chat_id": str(chat_id),
                "text": texto,
                "parse_mode": "HTML",
            },
            timeout=10,
        )

        if not response.ok:
            print("Error Telegram:", response.text)
            return False

        return True

    except Exception as e:
        print("Error enviando Telegram:", e)
        return False

def obtener_telegram_id_ti_por_nombre(nombre_ti):
    if not nombre_ti:
        return None

    ut = user_table()

    if not ut:
        return None

    cs = cols(ut)

    idc = find_col(cs, ["telegram_id", "chat_id", "id_telegram", "telegram", "user_id"])
    namec = find_col(cs, ["nombre", "name", "usuario", "username"])

    if not idc or not namec:
        return None

    try:
        with con() as c:
            row = c.execute(
                f"""
                select {q(idc)} as telegram_id
                from {q(ut)}
                where lower(cast({q(namec)} as text)) = lower(?)
                limit 1
                """,
                (nombre_ti,),
            ).fetchone()

            if row:
                return str(row["telegram_id"])

    except Exception as e:
        print("Error buscando Telegram ID del TI:", e)

    return None


def notificar_cambio_ticket(ticket_anterior, ticket_nuevo, cambios):
    ticket_id = ticket_nuevo.get("id", "")
    usuario = ticket_nuevo.get("usuario", "")
    area = ticket_nuevo.get("area", "")
    descripcion = ticket_nuevo.get("descripcion", "")
    estado = ticket_nuevo.get("estado", "")
    asignado = ticket_nuevo.get("asignado_a", "") or "Sin asignar"
    observacion = ticket_nuevo.get("observacion", "") or "Sin observación"

    # Notificación al usuario que creó el ticket
    chat_id_usuario = ticket_nuevo.get("chat_id")

    if chat_id_usuario:
        mensaje_usuario = [
            f"🔔 <b>Actualización de ticket #{ticket_id}</b>",
            "",
            f"<b>Estado:</b> {estado}",
            f"<b>TI asignado:</b> {asignado}",
        ]

        if "observacion" in cambios:
            mensaje_usuario.append(f"<b>Observación TI:</b> {observacion}")

        mensaje_usuario.append("")
        mensaje_usuario.append("Tu ticket fue actualizado por el equipo TI.")

        enviar_mensaje_telegram(chat_id_usuario, "\n".join(mensaje_usuario))

    # Notificación al TI asignado
    asignado_antes = str(ticket_anterior.get("asignado_a") or "").strip()
    asignado_despues = str(ticket_nuevo.get("asignado_a") or "").strip()

    if asignado_despues and asignado_despues != asignado_antes:
        telegram_id_ti = obtener_telegram_id_ti_por_nombre(asignado_despues)

        if telegram_id_ti:
            mensaje_ti = [
                f"📌 <b>Ticket asignado #{ticket_id}</b>",
                "",
                f"<b>Usuario:</b> {usuario}",
                f"<b>Área:</b> {area}",
                f"<b>Estado:</b> {estado}",
                "",
                "<b>Descripción:</b>",
                descripcion or "Sin descripción",
            ]

            if observacion and observacion != "Sin observación":
                mensaje_ti.extend([
                    "",
                    "<b>Observación TI:</b>",
                    observacion,
                ])

            enviar_mensaje_telegram(telegram_id_ti, "\n".join(mensaje_ti))

@app.route("/api/tickets/<int:rowid>", methods=["PUT"])
def api_ticket_update(rowid):
    tt = ticket_table()

    if not tt:
        return jsonify({"ok": False, "msg": "No existe tabla de tickets"}), 400

    data = request.json or {}
    cs = cols(tt)

    allowed_keys = [
        "asignado_a",
        "estado",
        "observacion",
    ]

    allowed = {k: v for k, v in data.items() if k in cs and k in allowed_keys}

    if not allowed:
        return jsonify(
            {"ok": False, "msg": "No hay campos válidos para actualizar"}
        ), 400

    try:
        with con() as c:
            ticket_actual_row = c.execute(
                f"select rowid as _rowid, * from {q(tt)} where rowid=?",
                (rowid,),
            ).fetchone()

            if not ticket_actual_row:
                return jsonify({"ok": False, "msg": "Ticket no encontrado"}), 404

            ticket_actual = dict(ticket_actual_row)

            estado_actual = str(ticket_actual.get("estado") or "").strip().lower()

            if estado_actual == "cerrado":
                return jsonify(
                    {
                        "ok": False,
                        "msg": "No se puede editar un ticket cerrado",
                    }
                ), 400

            nuevo_estado = str(
                allowed.get("estado", ticket_actual.get("estado") or "")
            ).strip()

            nuevo_asignado = str(
                allowed.get("asignado_a", ticket_actual.get("asignado_a") or "")
            ).strip()

            nueva_observacion = str(
                allowed.get("observacion", ticket_actual.get("observacion") or "")
            ).strip()

            if nuevo_estado in ["Proceso", "En proceso", "En Proceso", "Cerrado"] and not nuevo_asignado:
                return jsonify(
                    {
                        "ok": False,
                        "msg": "No puedes dejar sin asignar un ticket En proceso o Cerrado",
                    }
                ), 400

            if nuevo_estado == "Cerrado" and not nueva_observacion:
                return jsonify(
                    {
                        "ok": False,
                        "msg": "Para cerrar el ticket debes ingresar una observación",
                    }
                ), 400

            cambios = {}

            for key, nuevo_valor in allowed.items():
                valor_anterior = str(ticket_actual.get(key) or "")
                valor_nuevo = str(nuevo_valor or "")

                if valor_anterior != valor_nuevo:
                    cambios[key] = {
                        "antes": valor_anterior,
                        "despues": valor_nuevo,
                    }

            if "fecha_actualizacion" in cs:
                allowed["fecha_actualizacion"] = datetime.now().strftime(
                    "%Y-%m-%d %H:%M:%S.%f"
                )

            if not cambios:
                return jsonify({"ok": True, "msg": "Sin cambios"})

            sets = ", ".join(f"{q(k)}=?" for k in allowed)
            params = list(allowed.values()) + [rowid]

            c.execute(f"update {q(tt)} set {sets} where rowid=?", params)
            c.commit()

            ticket_nuevo_row = c.execute(
                f"select rowid as _rowid, * from {q(tt)} where rowid=?",
                (rowid,),
            ).fetchone()

            ticket_nuevo = dict(ticket_nuevo_row)

        notificar_cambio_ticket(ticket_actual, ticket_nuevo, cambios)

        try:
            from bot.services.sync_jobs_service import crear_job_sync

            crear_job_sync(ticket_nuevo.get("id"))
        except Exception as e:
            print("Error creando job sync desde panel:", e)

        return jsonify({"ok": True})

    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400


@app.route("/api/users", methods=["POST", "DELETE"])
def api_users():
    ut = user_table()

    if not ut:
        return jsonify({"ok": False, "msg": "No existe tabla de usuarios"}), 400

    cs = cols(ut)

    idc = find_col(cs, ["telegram_id", "chat_id", "id_telegram", "telegram", "user_id"])
    namec = find_col(cs, ["nombre", "name", "usuario", "username"])
    createdc = find_col(cs, ["creado", "created_at", "fecha_creacion"])

    if not idc:
        return jsonify({"ok": False, "msg": "No encuentro columna telegram_id/chat_id"}), 400

    try:
        with con() as c:
            if request.method == "POST":
                data = request.json or {}
                tid = data.get("telegram_id", "").strip()
                name = data.get("nombre", "").strip()

                if not tid:
                    return jsonify({"ok": False, "msg": "ID vacío"}), 400

                fields = [idc]
                values = [tid]

                if namec and name:
                    fields.append(namec)
                    values.append(name)

                if createdc:
                    fields.append(createdc)
                    values.append(datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f"))

                columns_sql = ", ".join(q(f) for f in fields)
                placeholders = ", ".join("?" for _ in fields)

                c.execute(
                    f"insert into {q(ut)} ({columns_sql}) values ({placeholders})",
                    values,
                )

                c.commit()
                return jsonify({"ok": True})

            data = request.json or {}
            tids = data.get("telegram_ids") or []

            if not tids:
                tid = data.get("telegram_id", "").strip()
                if tid:
                    tids = [tid]

            tids = [str(t).strip() for t in tids if str(t).strip()]

            if not tids:
                return jsonify({"ok": False, "msg": "No hay usuarios seleccionados"}), 400

            current_user = str(session.get("telegram_id", "")).strip()

            if current_user in tids:
                return jsonify({
                    "ok": False,
                    "msg": "No puedes eliminar el usuario con el que estás conectado"
                }), 400

            placeholders = ", ".join("?" for _ in tids)

            c.execute(
                f"delete from {q(ut)} where cast({q(idc)} as text) in ({placeholders})",
                tids,
            )

            c.commit()
            return jsonify({"ok": True, "deleted": len(tids)})

    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400


@app.get("/api/launcher/status")
def api_launcher_status():
    data = get_status()
    data["all"] = all(data.values()) if data else False
    return jsonify(data)


@app.post("/api/launcher/<service>/start")
def api_launcher_start(service):
    if service == "all":
        return jsonify(start_all())

    ok, msg = start_service(service)
    return jsonify({"ok": ok, "msg": msg}), 200 if ok else 400


@app.post("/api/launcher/<service>/stop")
def api_launcher_stop(service):
    if service == "all":
        return jsonify(stop_all())

    ok, msg = stop_service(service)
    return jsonify({"ok": ok, "msg": msg}), 200 if ok else 400


def build_ticket_excel(rows=None):
    template_path = BASE_DIR / "templates_excel" / "PlantillaReporte.xlsx"

    wb = load_workbook(template_path)
    ws = wb["Tickets"]

    export_columns = [
        "id",
        "usuario",
        "area",
        "descripcion",
        "estado",
        "fecha_creacion",
        "asignado_a",
        "observacion",
        "fecha_actualizacion",
    ]

    ws["A2"] = datetime.now().strftime("%d-%m-%Y %H:%M:%S")

    start_row = 4

    for row in range(start_row, ws.max_row + 1):
        for col in range(1, len(export_columns) + 1):
            ws.cell(row, col).value = None

    rows = rows or []

    for r_idx, row_data in enumerate(rows, start_row):
        for c_idx, key in enumerate(export_columns, 1):
            value = row_data.get(key, "")

            if key in ["fecha_creacion", "fecha_actualizacion"] and value:
                try:
                    value = datetime.fromisoformat(str(value)).strftime(
                        "%d-%m-%Y %H:%M:%S"
                    )
                except Exception:
                    pass

            ws.cell(r_idx, c_idx, value)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return bio


@app.post("/api/export/tickets")
def export_tickets():
    data = request.json or {}
    rows = data.get("rows", [])

    return send_file(
        build_ticket_excel(rows),
        as_attachment=True,
        download_name=f"reporte_tickets_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

@app.route("/api/email/tickets", methods=["POST"])
def email_tickets():
    to = os.getenv("EMAIL_TO", "").strip()
    smtp_host = os.getenv("SMTP_HOST", "").strip()
    smtp_user = os.getenv("SMTP_USER", "").strip()
    smtp_pass = os.getenv("SMTP_PASS", "").strip()
    smtp_port = int(os.getenv("SMTP_PORT", "587"))

    if not all([to, smtp_host, smtp_user, smtp_pass]):
        return jsonify(
            {
                "ok": False,
                "msg": "Configura EMAIL_TO, SMTP_HOST, SMTP_USER y SMTP_PASS en .env",
            }
        ), 400

    data_json = request.json or {}
    rows = data_json.get("rows", [])

    bio = build_ticket_excel(rows)
    data = bio.getvalue()

    msg = EmailMessage()
    msg["Subject"] = "Reporte de Tickets TicketIT"
    msg["From"] = smtp_user
    msg["To"] = to
    msg.set_content("Se adjunta reporte de tickets generado desde TicketIT.")
    msg.add_attachment(
        data,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="reporte_tickets.xlsx",
    )

    try:
        with smtplib.SMTP(smtp_host, smtp_port) as s:
            s.starttls()
            s.login(smtp_user, smtp_pass)
            s.send_message(msg)

        return jsonify({"ok": True})

    except Exception as e:
        return jsonify({"ok": False, "msg": str(e)}), 400

@app.route("/logs")
def logs():
    return render_template("logs.html")


@app.get("/api/logs")
def api_logs():
    log_dir = Path(os.getenv("LOG_DIR", str(BASE_DIR.parent / "logs")))

    files = {
        "worker": "worker.log",
        "webhook": "webhook.log",
        "ngrok": "ngrok.log",
        "panel": "panel.log",
    }

    data = {}

    for key, filename in files.items():
        path = log_dir / filename

        if path.exists():
            lines = path.read_text(
                encoding="utf-8",
                errors="replace"
            ).splitlines()

            data[key] = "\n".join(lines[-300:])
        else:
            data[key] = "Archivo de log no encontrado."

    return jsonify(data) 


if __name__ == "__main__":
    app.run(
        host=os.getenv("FLASK_HOST", os.getenv("HOST", "0.0.0.0")),
        port=int(os.getenv("FLASK_PORT", os.getenv("PORT", "8080"))),
        debug=os.getenv("DEBUG", "1") == "1",
    )