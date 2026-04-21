import smtplib
from datetime import datetime
from email.message import EmailMessage
from io import BytesIO


import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


from bot.config import DESTINATARIO, EMAIL_PASS, REMITENTE


#!---------------------------------------------------------
#! GENERA EXCEL
def generar_excel(tickets):
    df = construir_dataframe(tickets)
    df = ordenar_dataframe(df)

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Reporte", startrow=2)

        ws = writer.sheets["Reporte"]

        ws["A1"] = "Reporte De Tickets"
        ws["A2"] = f"Fecha de Reporte: {datetime.now().strftime('%d-%m-%Y (%H:%M:%S)')}"

        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].font = Font(size=10)

        aplicar_estilos_generales(ws)
        aplicar_estilo_headers(ws)
        auto_ajustar_columnas(ws)
        aplicar_filtros(ws)
        congelar_encabezado(ws)
        aplicar_colores_estado(ws)
        resaltar_nulos(ws)
        ajustar_columnas_especiales(ws)
        aplicar_bordes(ws)

    output.seek(0)
    return output


#!---------------------------------------------------------
#! CONSTRUYE LA ESTRUCTURA DE DATOS
def construir_dataframe(tickets):
    data = []

    for t in tickets:
        data.append(
            {
                "ID": t.id,
                "Usuario": t.usuario,
                "Área": t.area,
                "Descripción": t.descripcion,
                "Estado": t.estado,
                "Fecha Creacion": t.fecha_creacion.strftime("%d-%m-%Y %H:%M:%S")
                if t.fecha_creacion
                else "",
                "TI Asignado": t.asignado_a or "Sin asignar",
                "Observacion TI": t.observacion or "Sin observaciones",
                "Fecha Actualiz.": t.fecha_actualizacion.strftime("%d-%m-%Y %H:%M:%S")
                if t.fecha_actualizacion
                else "",
            }
        )

    return pd.DataFrame(data)


#!---------------------------------------------------------
#! ESTILOS GENERALES
def aplicar_estilos_generales(ws):
    for row in ws.iter_rows(min_row=4):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
    for col in ["D", "I"]:
        for cell in ws[col]:
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )


#!---------------------------------------------------------
#! ESTILO HEADERS
def aplicar_estilo_headers(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")

    for cell in ws[3]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].heigth = 20


#!---------------------------------------------------------
#! AJUSTAR COLUMNAS
def auto_ajustar_columnas(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = max_length + 2


#!---------------------------------------------------------
#! FILTROS
def aplicar_filtros(ws):
    ws.auto_filter.ref = f"A3:I{ws.max_row}"


# !CONGELAR ENCABEZADO
def congelar_encabezado(ws):
    ws.freeze_panes = "A4"


#! COLORES DE ESTADO
def aplicar_colores_estado(ws):
    for row in ws.iter_rows(min_row=4):
        estado = row[4].value

        if estado == "Abierto":
            fill = PatternFill(start_color="C6EFCE", fill_type="solid")

        elif estado == "Cerrado":
            fill = PatternFill(start_color="FFC7CE", fill_type="solid")

        elif estado == "En Proceso":
            fill = PatternFill(start_color="FFD966", fill_type="solid")

        else:
            continue

        for cell in row:
            cell.fill = fill


#! NULOS
def resaltar_nulos(ws):
    null_fill = PatternFill(start_color="EEECE1", fill_type="solid")

    for cell in ws["H"]:
        if cell.value == "Sin asignar":
            cell.fill = null_fill

    for cell in ws["I"]:
        if cell.value == "Sin Comentarios":
            cell.fill = null_fill


#!---------------------------------------------------------
# !COLUMNAS ESPECIALES
def ajustar_columnas_especiales(ws):
    tamaños = {
        "A": 7,
        "B": 12,
        "C": 12,
        "D": 40,
        "E": 10,
        "F": 20,
        "G": 12,
        "H": 40,
        "I": 20,
    }

    for col, width in tamaños.items():
        ws.column_dimensions[col].width = width


#!---------------------------------------------------------
#! ORDENAR
def ordenar_dataframe(df):
    return df.sort_values(by="ID", ascending=False)


def aplicar_bordes(ws):
    thin = Side(style="thin", color="000000")

    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.border = border


#!--------------------------------------------------------
#! ENVIAR REPORTE A CORREO
def enviar_report_correo(archivo_bytes, nombre_archivo):

    msg = EmailMessage()
    msg["Subject"] = f"Reporte Generado: {nombre_archivo}"
    msg["From"] = f"TI-BOT SOPORTE (No-Reply) <{REMITENTE}>"
    msg["To"] = DESTINATARIO
    msg.set_content(
        "Adjunto encontraras el reporte solicitado desde el Bot de Telegram TI-BOT."
    )

    payload = (
        archivo_bytes.getvalue()
        if hasattr(archivo_bytes, "getvalue")
        else archivo_bytes
    )

    msg.add_attachment(
        payload,
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=nombre_archivo,
    )
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(REMITENTE, EMAIL_PASS)
        smtp.send_message(msg)


#!--------------------------------------------------------

